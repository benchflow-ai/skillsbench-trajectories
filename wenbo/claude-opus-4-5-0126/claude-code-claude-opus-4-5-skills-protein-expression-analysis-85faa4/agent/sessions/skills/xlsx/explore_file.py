import pandas as pd
from openpyxl import load_workbook

wb = load_workbook('/root/protein_expression.xlsx')
print("Sheets:", wb.sheetnames)

# Explore Task sheet structure
task = wb['Task']
print("\n=== Task Sheet ===")
print(f"Dimensions: {task.dimensions}")
for row in range(1, 45):
    row_data = []
    for col in range(1, 15):
        val = task.cell(row=row, column=col).value
        if val is not None:
            row_data.append(f"{col}:{val}")
    if row_data:
        print(f"Row {row}: {row_data}")

# Explore Data sheet structure
data = wb['Data']
print("\n=== Data Sheet ===")
print(f"Dimensions: {data.dimensions}")
print("\nFirst row (headers):")
for col in range(1, 55):
    val = data.cell(row=1, column=col).value
    if val:
        print(f"  Col {col}: {val}")

print("\nFirst column (protein IDs), rows 1-10:")
for row in range(1, 11):
    print(f"  Row {row}: {data.cell(row=row, column=1).value}")
