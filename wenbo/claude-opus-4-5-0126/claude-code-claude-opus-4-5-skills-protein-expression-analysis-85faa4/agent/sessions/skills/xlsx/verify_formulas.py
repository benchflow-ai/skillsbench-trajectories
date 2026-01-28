from openpyxl import load_workbook

wb = load_workbook('/root/protein_expression.xlsx')  # Without data_only to see formulas
task = wb['Task']

print("=== Sample formulas in C11:L20 ===")
print(f"C11: {task.cell(row=11, column=3).value}")
print(f"L11: {task.cell(row=11, column=12).value}")
print(f"C20: {task.cell(row=20, column=3).value}")

print("\n=== Statistics formulas (row 24-27, col B) ===")
for row in range(24, 28):
    print(f"B{row}: {task.cell(row=row, column=2).value}")

print("\n=== Fold change formulas (rows 32-33) ===")
for row in range(32, 34):
    print(f"Row {row}:")
    print(f"  A: {task.cell(row=row, column=1).value}")
    print(f"  B: {task.cell(row=row, column=2).value}")
    print(f"  C (FC): {task.cell(row=row, column=3).value}")
    print(f"  D (Log2FC): {task.cell(row=row, column=4).value}")
