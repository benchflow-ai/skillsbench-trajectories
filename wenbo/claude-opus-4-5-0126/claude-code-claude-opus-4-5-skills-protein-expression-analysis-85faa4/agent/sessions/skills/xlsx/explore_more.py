from openpyxl import load_workbook

wb = load_workbook('/root/protein_expression.xlsx')
task = wb['Task']

print("=== Row 31-41 (Fold change section) ===")
for row in range(31, 42):
    row_data = []
    for col in range(1, 6):
        val = task.cell(row=row, column=col).value
        if val is not None:
            row_data.append(f"Col{col}:{val}")
    print(f"Row {row}: {row_data}")

print("\n=== Checking yellow cells (fills) ===")
for row in range(11, 21):
    for col in range(3, 13):
        cell = task.cell(row=row, column=col)
        fill = cell.fill
        if fill.start_color and fill.start_color.rgb:
            print(f"Cell ({row},{col}): fill={fill.start_color.rgb}")
            break
    break

print("\n=== Row 23-27 columns B-K ===")
for row in range(23, 28):
    row_data = []
    for col in range(1, 13):
        val = task.cell(row=row, column=col).value
        row_data.append(f"{col}:{val}")
    print(f"Row {row}: {row_data}")
