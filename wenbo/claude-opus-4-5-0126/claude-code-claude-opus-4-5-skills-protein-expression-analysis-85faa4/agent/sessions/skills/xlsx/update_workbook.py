from openpyxl import load_workbook

wb = load_workbook('/root/protein_expression.xlsx')
task = wb['Task']
data = wb['Data']

# Build mapping of sample names to Data sheet column letters
sample_to_col = {}
for col in range(4, 54):  # Data sheet has samples in columns D onwards
    sample_name = data.cell(row=1, column=col).value
    if sample_name:
        from openpyxl.utils import get_column_letter
        sample_to_col[sample_name] = get_column_letter(col)

# Build mapping of protein IDs to Data sheet row numbers
protein_to_row = {}
for row in range(2, 202):  # 200 proteins starting at row 2
    protein_id = data.cell(row=row, column=1).value
    if protein_id:
        protein_to_row[protein_id] = row

# Get sample names from Task sheet row 10 (columns C-L = 3-12)
task_samples = []
for col in range(3, 13):
    task_samples.append(task.cell(row=10, column=col).value)

# Get protein IDs from Task sheet rows 11-20, column A
task_proteins = []
for row in range(11, 21):
    task_proteins.append(task.cell(row=row, column=1).value)

print("Task samples:", task_samples)
print("Task proteins:", task_proteins)

# Step 1: Fill expression data C11:L20 with INDEX-MATCH formulas
# Formula: =INDEX(Data!$D$2:$BA$201, MATCH(A11, Data!$A$2:$A$201, 0), MATCH(C$10, Data!$D$1:$BA$1, 0))
for row_idx, protein_id in enumerate(task_proteins):
    task_row = 11 + row_idx
    for col_idx, sample_name in enumerate(task_samples):
        task_col = 3 + col_idx
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(task_col)
        # INDEX-MATCH formula for 2D lookup
        formula = f'=INDEX(Data!$D$2:$BA$201,MATCH($A{task_row},Data!$A$2:$A$201,0),MATCH({col_letter}$10,Data!$D$1:$BA$1,0))'
        task.cell(row=task_row, column=task_col).value = formula

# Step 2: Calculate statistics in rows 24-27, columns B-K
# Control samples are in columns C-G (columns 3-7 in Task sheet)
# Treated samples are in columns H-L (columns 8-12 in Task sheet)

# Get column letters for each protein's data
for col_idx in range(10):  # 10 proteins
    stat_col = 2 + col_idx  # B=2, C=3, ..., K=11
    data_row = 11 + col_idx  # Data for protein is in row 11+idx
    col_letter = get_column_letter(stat_col)

    # Control data is in C{data_row}:G{data_row}
    # Treated data is in H{data_row}:L{data_row}
    # Wait - the stats columns B-K should correspond to the 10 proteins
    # Let me re-check: proteins are in rows 11-20, and stats should be per protein

# Actually, looking at the layout again:
# Row 24-27 should have stats for each protein, with proteins as columns
# The yellow cells are in rows 24-27, columns B-K (10 columns for 10 proteins)

# For each protein (column B-K corresponds to proteins 1-10):
for protein_idx in range(10):
    stat_col = 2 + protein_idx  # B=2 through K=11
    data_row = 11 + protein_idx  # Expression data row for this protein
    col_letter = get_column_letter(stat_col)

    # Control Mean: average of columns C-G for this protein's row
    task.cell(row=24, column=stat_col).value = f'=AVERAGE($C{data_row}:$G{data_row})'
    # Control StdDev
    task.cell(row=25, column=stat_col).value = f'=STDEV($C{data_row}:$G{data_row})'
    # Treated Mean: average of columns H-L for this protein's row
    task.cell(row=26, column=stat_col).value = f'=AVERAGE($H{data_row}:$L{data_row})'
    # Treated StdDev
    task.cell(row=27, column=stat_col).value = f'=STDEV($H{data_row}:$L{data_row})'

# Step 3: Fold change calculations in rows 32-41, columns A-D
# Column A: Protein_ID, Column B: Gene_Symbol, Column C: Fold Change, Column D: Log2 FC
for protein_idx in range(10):
    fc_row = 32 + protein_idx  # Rows 32-41
    data_row = 11 + protein_idx  # Expression data row
    stat_col = 2 + protein_idx  # Statistics column (B-K)
    stat_col_letter = get_column_letter(stat_col)

    # Copy Protein_ID and Gene_Symbol from rows 11-20
    task.cell(row=fc_row, column=1).value = f'=$A{data_row}'
    task.cell(row=fc_row, column=2).value = f'=$B{data_row}'

    # Log2 FC = Treated Mean - Control Mean (row 26 - row 24)
    task.cell(row=fc_row, column=4).value = f'=${stat_col_letter}$26-${stat_col_letter}$24'

    # Fold Change = 2^(Log2 FC)
    task.cell(row=fc_row, column=3).value = f'=POWER(2,D{fc_row})'

wb.save('/root/protein_expression.xlsx')
print("Workbook saved successfully!")
