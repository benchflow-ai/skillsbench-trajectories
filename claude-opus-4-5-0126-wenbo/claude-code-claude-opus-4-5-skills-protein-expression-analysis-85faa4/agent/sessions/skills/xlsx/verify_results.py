from openpyxl import load_workbook

wb = load_workbook('/root/protein_expression.xlsx', data_only=True)
task = wb['Task']

print("=== Expression Data (C11:L20) - first 3 rows ===")
for row in range(11, 14):
    row_data = []
    for col in range(3, 13):
        val = task.cell(row=row, column=col).value
        row_data.append(f"{val:.3f}" if val else "None")
    print(f"Row {row}: {row_data}")

print("\n=== Statistics (rows 24-27, cols B-K) ===")
labels = ['Control Mean', 'Control StdDev', 'Treated Mean', 'Treated StdDev']
for row, label in zip(range(24, 28), labels):
    row_data = []
    for col in range(2, 12):
        val = task.cell(row=row, column=col).value
        row_data.append(f"{val:.3f}" if val else "None")
    print(f"{label}: {row_data}")

print("\n=== Fold Change (rows 32-41) ===")
print("Protein_ID | Gene | FC | Log2FC")
for row in range(32, 42):
    protein = task.cell(row=row, column=1).value
    gene = task.cell(row=row, column=2).value
    fc = task.cell(row=row, column=3).value
    log2fc = task.cell(row=row, column=4).value
    if protein and fc:
        print(f"{protein[:25]:25} | {gene:8} | {fc:.4f} | {log2fc:.4f}")
